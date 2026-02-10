import pygame
import serial
import random
import subprocess
import tkinter as tk
from PIL import Image, ImageTk, ImageSequence
import os
import sys
import threading
import time
import numpy as np
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

# MediaPipe for attention tracking
import cv2
import mediapipe as mp

print("=== Level 2: Image Identification Mode Started ===")

SERIAL_PORT = '/dev/ttyUSB0'
BAUD_RATE = 9600

BASE_DIR = "/home/pi/asd_learning_system"
IMAGE_PATH = os.path.join(BASE_DIR, "animal_images")
SOUND_PATH = os.path.join(BASE_DIR, "animal_sounds")
GIF_PATH = os.path.join(BASE_DIR, "feedback_gifs")
RESULTS_FILE = os.path.join(BASE_DIR, "level2_results.xlsx")
ATTENTION_FILE = os.path.join(BASE_DIR, "level2_attention_scores.xlsx")

TOTAL_QUESTIONS = 14
MAX_RETRIES = 3
LOOKDOWN_THRESHOLD = 3.0  # seconds allowed to look down for RFID card

gif_running = False
current_gif_frames = []
accepting_input = False

# Get name from command line or show entry screen
child_name = sys.argv[1] if len(sys.argv) > 1 else ""

animal_data = {
    "936FA320": "cat",
    "33719E20": "horse",
    "C376B420": "lion",
    "536E4BE4": "parrot",
    "33A8B920": "dog",
    "339FFC30": "sheep",
    "03548F22": "donkey",
    "131CF430": "monkey",
    "236A8222": "duck",
    "135C7A31": "crow",
    "73C28131": "cow",
    "E3257622": "dolphin",
    "23EB9022": "frog",
    "033D9423": "chicken"
}

animal_to_uid = {v: k for k, v in animal_data.items()}
animal_list = list(animal_to_uid.keys())
shuffled_animals = []

try:
    ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1)
except:
    print("Serial error")
    sys.exit()

pygame.init()
pygame.mixer.init()

# ========== ATTENTION TRACKING ==========

# MediaPipe Face Mesh
mp_face_mesh = mp.solutions.face_mesh
face_mesh = mp_face_mesh.FaceMesh(
    max_num_faces=1,
    refine_landmarks=True,
    min_detection_confidence=0.5,
    min_tracking_confidence=0.5
)

# Camera
camera_cap = None
camera_thread = None
camera_running = False

# Attention state
face_detected = False
head_attentive = False
gaze_attentive = False
attention_active = False

# Timing for current animal
current_animal_for_attention = None
stimulus_start_time = None
attention_duration = 0.0
lookdown_start_time = None
last_attention_true_time = None

# Attention scores for all animals
attention_scores = []

# Thread lock
attention_lock = threading.Lock()

def rotation_matrix_to_euler_angles(R):
    """Convert rotation matrix to Euler angles"""
    sy = np.sqrt(R[0, 0] * R[0, 0] + R[1, 0] * R[1, 0])
    singular = sy < 1e-6
    
    if not singular:
        x = np.arctan2(R[2, 1], R[2, 2])
        y = np.arctan2(-R[2, 0], sy)
        z = np.arctan2(R[1, 0], R[0, 0])
    else:
        x = np.arctan2(-R[1, 2], R[1, 1])
        y = np.arctan2(-R[2, 0], sy)
        z = 0
    
    return np.degrees([x, y, z])

def check_head_pose(landmarks, frame_shape):
    """Check if head is facing forward (|yaw| ≤ 25°, |pitch| ≤ 20°)"""
    h, w = frame_shape[:2]
    
    landmark_indices = [1, 152, 33, 263, 61, 291]
    
    model_points = np.array([
        (0.0, 0.0, 0.0),
        (0.0, -330.0, -65.0),
        (-225.0, 170.0, -135.0),
        (225.0, 170.0, -135.0),
        (-150.0, -150.0, -125.0),
        (150.0, -150.0, -125.0)
    ], dtype=np.float64)
    
    image_points = np.array([
        (landmarks.landmark[idx].x * w, landmarks.landmark[idx].y * h)
        for idx in landmark_indices
    ], dtype=np.float64)
    
    focal_length = w
    center = (w / 2, h / 2)
    camera_matrix = np.array([
        [focal_length, 0, center[0]],
        [0, focal_length, center[1]],
        [0, 0, 1]
    ], dtype=np.float64)
    
    dist_coeffs = np.zeros((4, 1))
    
    success, rotation_vector, _ = cv2.solvePnP(
        model_points, 
        image_points, 
        camera_matrix, 
        dist_coeffs,
        flags=cv2.SOLVEPNP_ITERATIVE
    )
    
    if not success:
        return False
    
    rotation_matrix, _ = cv2.Rodrigues(rotation_vector)
    pitch, yaw, roll = rotation_matrix_to_euler_angles(rotation_matrix)
    
    return abs(yaw) <= 25 and abs(pitch) <= 20

def check_eye_gaze(landmarks):
    """Check if eyes are looking at screen (iris centered)"""
    try:
        # Get iris and eye corner landmarks
        left_iris = (landmarks.landmark[468].x, landmarks.landmark[468].y)
        left_outer = (landmarks.landmark[33].x, landmarks.landmark[33].y)
        left_inner = (landmarks.landmark[133].x, landmarks.landmark[133].y)
        
        right_iris = (landmarks.landmark[473].x, landmarks.landmark[473].y)
        right_outer = (landmarks.landmark[263].x, landmarks.landmark[263].y)
        right_inner = (landmarks.landmark[362].x, landmarks.landmark[362].y)
        
        # Check if iris is centered (0.3 to 0.7 range)
        left_eye_width = abs(left_outer[0] - left_inner[0])
        right_eye_width = abs(right_outer[0] - right_inner[0])
        
        left_pos = (left_iris[0] - left_inner[0]) / left_eye_width if left_eye_width > 0 else 0.5
        right_pos = (right_iris[0] - right_inner[0]) / right_eye_width if right_eye_width > 0 else 0.5
        
        left_centered = 0.3 <= left_pos <= 0.7
        right_centered = 0.3 <= right_pos <= 0.7
        
        return left_centered and right_centered
    except:
        return False

def attention_tracking_loop():
    """Background thread that tracks attention"""
    global camera_running, face_detected, head_attentive, gaze_attentive
    global attention_active, attention_duration, lookdown_start_time
    global last_attention_true_time, stimulus_start_time
    
    print("[ATTENTION] Tracking thread started")
    
    while camera_running:
        ret, frame = camera_cap.read()
        if not ret:
            time.sleep(0.033)
            continue
        
        # Process with MediaPipe
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        results = face_mesh.process(rgb_frame)
        
        # Check criteria
        face_ok = False
        head_ok = False
        gaze_ok = False
        
        if results.multi_face_landmarks:
            landmarks = results.multi_face_landmarks[0]
            face_ok = True
            head_ok = check_head_pose(landmarks, frame.shape)
            gaze_ok = check_eye_gaze(landmarks)
        
        current_time = time.time()
        
        with attention_lock:
            face_detected = face_ok
            head_attentive = head_ok
            gaze_attentive = gaze_ok
            
            all_criteria_met = face_ok and head_ok and gaze_ok
            
            # Only track if we have an active animal
            if current_animal_for_attention is not None and stimulus_start_time is not None:
                
                # ATTENTION ACTIVE
                if all_criteria_met:
                    if not attention_active:
                        print(f"[ATTENTION] ACTIVE (Face:{face_ok}, Head:{head_ok}, Gaze:{gaze_ok})")
                        attention_active = True
                        last_attention_true_time = current_time
                        lookdown_start_time = None
                    else:
                        # Continue counting attention time
                        time_delta = current_time - last_attention_true_time
                        attention_duration += time_delta
                        last_attention_true_time = current_time
                
                # NOT ATTENTIVE
                else:
                    if attention_active:
                        print(f"[ATTENTION] INACTIVE (Face:{face_ok}, Head:{head_ok}, Gaze:{gaze_ok})")
                        attention_active = False
                        
                        # Start lookdown timer
                        if lookdown_start_time is None:
                            lookdown_start_time = current_time
                            print(f"[ATTENTION] Lookdown timer started (threshold: {LOOKDOWN_THRESHOLD}s)")
        
        time.sleep(0.033)  # ~30 FPS
    
    print("[ATTENTION] Tracking thread stopped")

def start_camera():
    """Start camera and tracking thread"""
    global camera_cap, camera_running, camera_thread
    
    camera_cap = cv2.VideoCapture(0)
    if not camera_cap.isOpened():
        print("[ATTENTION] ERROR: Camera failed to open!")
        return False
    
    camera_cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
    camera_cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    camera_cap.set(cv2.CAP_PROP_FPS, 30)
    
    camera_running = True
    camera_thread = threading.Thread(target=attention_tracking_loop, daemon=True)
    camera_thread.start()
    
    print("[ATTENTION] Camera and tracking started")
    return True

def stop_camera():
    """Stop camera and tracking"""
    global camera_running, camera_cap, camera_thread
    
    camera_running = False
    if camera_thread:
        camera_thread.join(timeout=2)
    if camera_cap:
        camera_cap.release()
    
    print("[ATTENTION] Camera stopped")

def start_tracking_animal(animal_name):
    """Start tracking attention for this animal"""
    global current_animal_for_attention, stimulus_start_time, attention_duration
    global lookdown_start_time, last_attention_true_time, attention_active
    
    with attention_lock:
        current_animal_for_attention = animal_name
        stimulus_start_time = time.time()
        attention_duration = 0.0
        lookdown_start_time = None
        last_attention_true_time = None
        attention_active = False
    
    print(f"[ATTENTION] Started tracking: {animal_name}")

def on_rfid_scan():
    """Called when RFID card is scanned"""
    global lookdown_start_time, attention_duration
    
    with attention_lock:
        if lookdown_start_time is not None:
            lookdown_time = time.time() - lookdown_start_time
            
            if lookdown_time <= LOOKDOWN_THRESHOLD:
                # Within threshold - count as attention
                attention_duration += lookdown_time
                print(f"[ATTENTION] RFID scanned in {lookdown_time:.2f}s - COUNTED")
            else:
                print(f"[ATTENTION] RFID scanned after {lookdown_time:.2f}s - OFF TASK")
            
            lookdown_start_time = None

def get_attention_score():
    """Get final attention score for current animal"""
    with attention_lock:
        if stimulus_start_time is None:
            return 0.0, 0.0, 0.0
        
        total_time = time.time() - stimulus_start_time
        
        if total_time > 0:
            percentage = (attention_duration / total_time) * 100
        else:
            percentage = 0.0
        
        return percentage, attention_duration, total_time

def save_attention_score(animal_name):
    """Save attention score for this animal"""
    percentage, attn_dur, total_dur = get_attention_score()
    
    attention_scores.append({
        'animal': animal_name,
        'percentage': round(percentage, 2),
        'attention_duration': round(attn_dur, 2),
        'total_duration': round(total_dur, 2),
        'timestamp': datetime.now()
    })
    
    print(f"[LOGGER] {animal_name}: {percentage:.1f}% ({attn_dur:.1f}s / {total_dur:.1f}s)")

def save_attention_to_excel():
    """Save all attention scores to Excel"""
    if len(attention_scores) == 0:
        print("[LOGGER] No attention data to save")
        return
    
    try:
        if os.path.exists(ATTENTION_FILE):
            wb = load_workbook(ATTENTION_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Attention Scores"
            
            headers = ["Child Name", "Animal", "Attention Score %", 
                      "Attention Duration (s)", "Total Duration (s)", "Date", "Time"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
        
        for score in attention_scores:
            date_str = score['timestamp'].strftime("%Y-%m-%d")
            time_str = score['timestamp'].strftime("%H:%M:%S")
            
            row = [
                child_name,
                score['animal'],
                score['percentage'],
                score['attention_duration'],
                score['total_duration'],
                date_str,
                time_str
            ]
            
            ws.append(row)
            
            row_num = ws.max_row
            for cell in ws[row_num]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(ATTENTION_FILE)
        print(f"[LOGGER] Saved {len(attention_scores)} scores to {ATTENTION_FILE}")
        
        # Print summary
        avg = sum(s['percentage'] for s in attention_scores) / len(attention_scores)
        best = max(attention_scores, key=lambda x: x['percentage'])
        worst = min(attention_scores, key=lambda x: x['percentage'])
        print(f"[SUMMARY] Average: {avg:.1f}% | Best: {best['animal']} ({best['percentage']:.1f}%) | Worst: {worst['animal']} ({worst['percentage']:.1f}%)")
        
    except Exception as e:
        print(f"[LOGGER] Error: {e}")

# ========== GAME CODE ==========

def exit_app(event=None):
    print("Exiting application safely...")
    global accepting_input
    accepting_input = False
    pygame.mixer.music.stop()
    
    # Stop camera and save attention data
    stop_camera()
    save_attention_to_excel()
    
    if ser.is_open:
        ser.close()
    root.quit()
    root.destroy()

root = tk.Tk()
root.attributes("-fullscreen", True)
root.bind("<Escape>", exit_app)
root.bind("<Control-q>", exit_app)
#root.overrideredirect(True)
root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")
root.configure(bg="#F4FFDB")
root.focus_force()

current_animal = None
question_count = 0
retry_count = 0
score = 0

# Main container
main_frame = tk.Frame(root, bg="#F4FFDB")
main_frame.pack(expand=True, fill="both")

image_label = tk.Label(main_frame, bg="#F4FFDB")
image_label.pack(expand=True)

status_label = tk.Label(main_frame, font=("Arial", 28), bg="#F4FFDB")
status_label.pack()

# ---------- UTILITIES ----------

def save_results_to_excel():
    """Save the game results to Excel file"""
    try:
        if os.path.exists(RESULTS_FILE):
            wb = load_workbook(RESULTS_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Level 2 Results"
            
            headers = ["Name", "Score", "Total Questions", "Percentage", "Date", "Time"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="457B9D", end_color="457B9D", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
        
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        percentage = round((score / TOTAL_QUESTIONS) * 100, 1)
        
        new_row = [child_name, score, TOTAL_QUESTIONS, f"{percentage}%", date_str, time_str]
        ws.append(new_row)
        
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(RESULTS_FILE)
        print(f"Results saved: {child_name} - {score}/{TOTAL_QUESTIONS} ({percentage}%)")
        return True
    except Exception as e:
        print(f"Error saving results: {e}")
        return False

def play_audio_non_blocking(file):
    """Play audio in a separate thread"""
    def play():
        try:
            pygame.mixer.music.load(file)
            pygame.mixer.music.play()
        except:
            pass
    threading.Thread(target=play, daemon=True).start()

def play_audio_blocking(file):
    try:
        pygame.mixer.music.load(file)
        pygame.mixer.music.play()
        while pygame.mixer.music.get_busy():
            root.update()
    except:
        pass

def show_current_animal():
    try:
        img = Image.open(os.path.join(IMAGE_PATH, f"{current_animal}_image.jpg"))
        img = img.resize((root.winfo_screenwidth(), root.winfo_screenheight()))
        photo = ImageTk.PhotoImage(img)
        image_label.config(image=photo)
        image_label.image = photo
    except:
        pass

def show_gif_with_audio(gif_name, audio_file=None, duration=2000, loop=False):
    """Show GIF and play audio simultaneously"""
    global gif_running, current_gif_frames
    gif_running = False
    root.update()
    gif_running = True

    try:
        gif_path = os.path.join(GIF_PATH, gif_name)
        gif_image = Image.open(gif_path)
        current_gif_frames = [
            ImageTk.PhotoImage(frame.resize(
                (root.winfo_screenwidth(), root.winfo_screenheight())))
            for frame in ImageSequence.Iterator(gif_image)
        ]

        if audio_file:
            play_audio_non_blocking(audio_file)

        def animate(idx=0):
            if not gif_running:
                return
            
            if not loop and idx >= len(current_gif_frames):
                if current_gif_frames:
                    image_label.config(image=current_gif_frames[-1])
                return
                
            image_label.config(image=current_gif_frames[idx % len(current_gif_frames)])
            root.after(100, animate, idx + 1)

        animate()
        
        if not loop:
            root.after(duration, stop_gif)
    except Exception as e:
        print(f"Error showing GIF: {e}")

def stop_gif():
    global gif_running
    gif_running = False

# ---------- START SCREEN ----------

def show_start_screen():
    """Display welcome screen"""
    global child_name, accepting_input
    
    accepting_input = False
    
    for widget in main_frame.winfo_children():
        widget.pack_forget()
    
    start_container = tk.Frame(main_frame, bg="#F4FFDB")
    start_container.pack(expand=True)
    
    gif_label = tk.Label(start_container, bg="#F4FFDB")
    gif_label.pack(pady=20)
    
    try:
        gif_path = os.path.join(GIF_PATH, "default_level2.gif")
        gif_image = Image.open(gif_path)
        gif_frames = [
            ImageTk.PhotoImage(frame.resize((600, 400)))
            for frame in ImageSequence.Iterator(gif_image)
        ]
        
        frame_count = [0]
        
        def animate_start_gif():
            if frame_count[0] < len(gif_frames) and gif_label.winfo_exists():
                gif_label.config(image=gif_frames[frame_count[0]])
                gif_label.image = gif_frames[frame_count[0]]
                frame_count[0] += 1
                root.after(100, animate_start_gif)
            elif gif_label.winfo_exists():
                gif_label.config(image=gif_frames[-1])
                gif_label.image = gif_frames[-1]
        
        animate_start_gif()
    except Exception as e:
        print(f"Error loading start GIF: {e}")
    
    name_frame = tk.Frame(start_container, bg="#F4FFDB")
    name_frame.pack(pady=30)
    
    name_label = tk.Label(
        name_frame, 
        text="Enter Your Name:", 
        font=("Arial", 32, "bold"),
        bg="#F4FFDB",
        fg="#1d3557"
    )
    name_label.pack(pady=10)
    
    name_entry = tk.Entry(
        name_frame,
        font=("Arial", 28),
        width=20,
        justify="center",
        bg="white",
        fg="#1d3557",
        relief="solid",
        bd=2
    )
    name_entry.pack(pady=10)
    name_entry.focus()
    
    def start_game_clicked():
        global child_name
        child_name = name_entry.get().strip()
        if not child_name:
            child_name = "Player"
        print(f"Starting game for: {child_name}")
        start_container.destroy()
        start_game()
    
    name_entry.bind("<Return>", lambda e: start_game_clicked())
    
    start_btn = tk.Button(
        name_frame,
        text="START",
        font=("Arial", 32, "bold"),
        bg="#457b9d",
        fg="white",
        activebackground="#1d3557",
        activeforeground="white",
        relief="raised",
        bd=5,
        padx=40,
        pady=15,
        command=start_game_clicked,
        cursor="hand2"
    )
    start_btn.pack(pady=20)

# ---------- GAME FLOW ----------

def start_game():
    global accepting_input, shuffled_animals
    
    # Create shuffled list of all animals
    shuffled_animals = animal_list.copy()
    random.shuffle(shuffled_animals)
    
    # Start camera and attention tracking
    if not start_camera():
        print("[WARNING] Game starting without attention tracking!")
    
    image_label.pack(expand=True)
    status_label.pack()
    
    accepting_input = False
    show_new_animal()

def show_new_animal():
    global current_animal, retry_count, question_count, accepting_input

    if question_count >= TOTAL_QUESTIONS:
        show_final_score()
        return

    stop_gif()
    retry_count = 0
    
    current_animal = shuffled_animals[question_count]
    question_count += 1

    print(f"Question {question_count}: {current_animal}")
    
    # Start attention tracking for this animal
    start_tracking_animal(current_animal)

    try:
        img = Image.open(os.path.join(IMAGE_PATH, f"{current_animal}_image.jpg"))
        img = img.resize((root.winfo_screenwidth(), root.winfo_screenheight()))
        photo = ImageTk.PhotoImage(img)
        image_label.config(image=photo)
        image_label.image = photo
        status_label.config(text=f"Question {question_count}/{TOTAL_QUESTIONS}", fg="#1d3557")
        
        accepting_input = True
    except Exception as e:
        print(f"Error showing animal: {e}")

def check_rfid():
    global retry_count, score, accepting_input

    if ser.in_waiting:
        uid = ser.readline().decode(errors="ignore").strip().upper()
        
        if not uid or not accepting_input or current_animal is None:
            root.after(100, check_rfid)
            return
            
        print(f"RFID scanned: {uid}")
        
        accepting_input = False
        
        # Notify attention tracker
        on_rfid_scan()

        if uid == animal_to_uid[current_animal]:
            score += 1
            
            # Save attention score
            save_attention_score(current_animal)
            
            play_audio_blocking(os.path.join(SOUND_PATH, f"{current_animal}_sound.mp3"))
            show_gif_with_audio(
                "goodJob.gif",
                os.path.join(SOUND_PATH, "correct_sound.mp3"),
                duration=2000
            )
            root.after(2500, show_new_animal)

        else:
            retry_count += 1
            if retry_count < MAX_RETRIES:
                show_gif_with_audio(
                    "tryAgain.gif",
                    os.path.join(SOUND_PATH, "incorrect_sound.mp3"),
                    duration=1500
                )
                
                def restore_after_retry():
                    show_current_animal()
                    global accepting_input
                    accepting_input = True
                
                root.after(1600, restore_after_retry)
            else:
                # Save attention score
                save_attention_score(current_animal)
                
                show_gif_with_audio(
                    "uhOh.gif",
                    os.path.join(SOUND_PATH, "oops_sound.mp3"),
                    duration=2000
                )
                root.after(2500, show_new_animal)
    
    root.after(100, check_rfid)

def show_final_score():
    """Display final score"""
    global gif_running, current_gif_frames, accepting_input
    
    accepting_input = False
    stop_gif()
    
    # Stop camera and save all attention data
    stop_camera()
    save_attention_to_excel()
    save_results_to_excel()
    
    status_label.pack_forget()
    image_label.pack_forget()
    
    final_container = tk.Frame(main_frame, bg="#F4FFDB")
    final_container.pack(expand=True)
    
    final_gif_label = tk.Label(final_container, bg="#F4FFDB")
    final_gif_label.pack(pady=20)
    
    score_text = f"Great Job, {child_name}!\n\nYou identified {score} out of {TOTAL_QUESTIONS} animals correctly!"
    score_label = tk.Label(
        final_container,
        text=score_text,
        font=("Arial", 36, "bold"),
        bg="#F4FFDB",
        fg="#1d3557",
        justify="center"
    )
    score_label.pack(pady=30)
    
    try:
        gif_path = os.path.join(GIF_PATH, "finalScore.gif")
        gif_image = Image.open(gif_path)
        final_frames = [
            ImageTk.PhotoImage(frame.resize((600, 400)))
            for frame in ImageSequence.Iterator(gif_image)
        ]
        
        frame_count = [0]
        
        def animate_final_gif():
            if frame_count[0] < len(final_frames) and final_gif_label.winfo_exists():
                final_gif_label.config(image=final_frames[frame_count[0]])
                final_gif_label.image = final_frames[frame_count[0]]
                frame_count[0] += 1
                root.after(100, animate_final_gif)
            elif final_gif_label.winfo_exists():
                final_gif_label.config(image=final_frames[-1])
                final_gif_label.image = final_frames[-1]
        
        animate_final_gif()
    except Exception as e:
        print(f"Error loading final GIF: {e}")
    
    exit_btn = tk.Button(
        final_container,
        text="EXIT",
        font=("Arial", 28, "bold"),
        bg="#e63946",
        fg="white",
        activebackground="#d62828",
        activeforeground="white",
        relief="raised",
        bd=5,
        padx=30,
        pady=10,
        command=exit_app,
        cursor="hand2"
    )
    exit_btn.pack(pady=20)

# ---------- START ----------
if child_name:
    print(f"Level 2 started for: {child_name}")
    start_game()
else:
    show_start_screen()

check_rfid()
root.mainloop()
