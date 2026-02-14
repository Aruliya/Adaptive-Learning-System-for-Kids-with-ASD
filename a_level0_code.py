import pygame
import subprocess
import tkinter as tk
from PIL import Image, ImageTk, ImageSequence
import os
import sys
import threading
import time
import cv2
import mediapipe as mp
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

print("=== Level 0: Continuous Learning with Visual Attention Tracking ===")

BASE_DIR = "/home/pi/asd_learning_system"
IMAGE_PATH = os.path.join(BASE_DIR, "animal_images")
SOUND_PATH = os.path.join(BASE_DIR, "animal_sounds")
GIF_PATH = os.path.join(BASE_DIR, "feedback_gifs")
ATTENTION_FILE = os.path.join(BASE_DIR, "level0_attention_tracking.xlsx")

# Get name from command line (or show entry screen later)
child_name = sys.argv[1] if len(sys.argv) > 1 else ""
if child_name:
    print(f"Level 0 started for: {child_name}")
else:
    print("Level 0: Waiting for name entry...")

# Animal data
animal_data = {
    "936FA320": "cat",
    "33719E20": "horse",
    "C376B420": "lion",
    "536E4BE4": "parrot",
    "33A8B920": "dog",
    "339FFC30": "sheep",
    "03548F22": "donkey",
    "131CF430": "monkey",
    "236A8222": "duck",
    "135C7A31": "crow",
    "73C28131": "cow",
    "E3257622": "dolphin",
    "23EB9022": "frog",
    "033D9423": "chicken"
}

animal_list = list(animal_data.values())
current_animal_idx = 0

# Audio playback
pygame.init()
pygame.mixer.init()

# Computer Vision setup
mp_face_detection = mp.solutions.face_detection
mp_face_mesh = mp.solutions.face_mesh
mp_drawing = mp.solutions.drawing_utils

face_detection = None
face_mesh = None
camera = None
camera_running = False

# Attention tracking metrics
ATTENTION_DURATION = 0  # Total time attention was true (seconds)
ATTENTION_SWITCHES = 0  # Number of times attention changed to True
attention_history = []  # List of (timestamp, is_attention) for analysis
session_start_time = 0
game_running = False

# Thresholds for attention estimation
FACE_DETECTION_CONFIDENCE = 0.5
HEAD_POSE_THRESHOLD = 25  # degrees - max head tilt allowed
EYE_GAZE_THRESHOLD = 0.3  # distance from center - lower is better

# GUI elements
root = None
image_label = None
status_label = None
attention_indicator = None

def play_sound(animal):
    """Play animal sound"""
    try:
        sound_file = os.path.join(SOUND_PATH, f"{animal}_sound.mp3")
        pygame.mixer.music.load(sound_file)
        pygame.mixer.music.play()
    except Exception as e:
        print(f"Error playing sound: {e}")

def show_animal(animal):
    """Display animal image"""
    try:
        img_path = os.path.join(IMAGE_PATH, f"{animal}_image.jpg")
        img = Image.open(img_path).resize(
            (root.winfo_screenwidth(), int(root.winfo_screenheight() * 0.85))
        )
        photo = ImageTk.PhotoImage(img)
        image_label.config(image=photo)
        image_label.image = photo
        
        status_label.config(text=f"{animal.upper()}", fg="#1d3557")
        print(f"Displaying: {animal}")
        
        # Play sound
        play_sound(animal)
    except Exception as e:
        print(f"Error showing animal: {e}")

def initialize_camera():
    """Initialize camera for face detection"""
    global camera, camera_running, face_detection, face_mesh
    try:
        for idx in range(0, 4):
            cam = cv2.VideoCapture(idx)
            if cam is not None and cam.isOpened():
                camera = cam
                break

        if camera is None or not camera.isOpened():
            print("✗ Failed to open camera")
            return False

        camera.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

        face_detection = mp_face_detection.FaceDetection(
            model_selection=0,
            min_detection_confidence=FACE_DETECTION_CONFIDENCE
        )

        face_mesh = mp_face_mesh.FaceMesh(
            static_image_mode=False,
            max_num_faces=1,
            min_detection_confidence=0.5,
            min_tracking_confidence=0.5
        )

        camera_running = True
        print("✓ Camera initialized for attention tracking")
        return True
    except Exception as e:
        print(f"✗ Camera error: {e}")
        return False

def estimate_head_pose(landmarks, image_width, image_height):
    """Estimate head pose from face mesh landmarks"""
    try:
        # Get reference points for head pose estimation
        nose = landmarks[1]  # Nose tip
        left_eye = landmarks[33]  # Left eye
        right_eye = landmarks[263]  # Right eye
        
        # Calculate angle based on eye positions
        eye_distance = abs(right_eye.x - left_eye.x)
        nose_x_offset = abs(nose.x - (left_eye.x + right_eye.x) / 2)
        
        # Estimate head rotation angle
        if eye_distance > 0:
            head_angle = np.arctan2(nose_x_offset, eye_distance) * (180 / np.pi)
        else:
            head_angle = 0
        
        return head_angle
    except Exception as e:
        return 90  # Return high angle on error (ignore)

def estimate_eye_gaze(landmarks):
    """Estimate if eyes are looking at screen center"""
    try:
        # Get iris points
        left_iris = landmarks[468]  # Left iris center
        right_iris = landmarks[473]  # Right iris center
        
        # Calculate average gaze direction
        avg_iris_x = (left_iris.x + right_iris.x) / 2
        avg_iris_y = (left_iris.y + right_iris.y) / 2
        
        # Calculate distance from center (0.5, 0.5 is center)
        gaze_distance = np.sqrt((avg_iris_x - 0.5) ** 2 + (avg_iris_y - 0.5) ** 2)
        
        return gaze_distance
    except Exception as e:
        return 1.0  # Return high distance on error (ignore)

def check_attention():
    """Check if child is paying attention based on 3 criteria"""
    if not camera_running or camera is None:
        return False
    
    try:
        ret, frame = camera.read()
        if not ret:
            return False
        
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        rgb_frame.flags.writeable = False
        
        # 1. Face Detection
        face_results = face_detection.process(rgb_frame)
        face_detected = face_results.detections is not None and len(face_results.detections) > 0
        
        # 2. Head Pose Estimation
        mesh_results = face_mesh.process(rgb_frame)
        head_pose_ok = False
        
        # 3. Eye Gaze Tracking
        gaze_ok = False
        
        if mesh_results.multi_face_landmarks and len(mesh_results.multi_face_landmarks) > 0:
            landmarks = mesh_results.multi_face_landmarks[0].landmark
            
            # Check head pose
            head_angle = estimate_head_pose(landmarks, frame.shape[1], frame.shape[0])
            head_pose_ok = abs(head_angle) < HEAD_POSE_THRESHOLD
            
            # Check eye gaze
            gaze_distance = estimate_eye_gaze(landmarks)
            gaze_ok = gaze_distance < EYE_GAZE_THRESHOLD
        
        # Attention is TRUE only if ALL three conditions are met
        attention = face_detected and head_pose_ok and gaze_ok
        
        return attention
    except Exception as e:
        print(f"Attention check error: {e}")
        return False

def attention_tracking_thread():
    """Run continuous attention tracking in background"""
    global ATTENTION_DURATION, ATTENTION_SWITCHES, attention_history, session_start_time
    
    last_attention_state = False
    last_update_time = time.time()
    
    while game_running:
        try:
            current_time = time.time()
            time_delta = current_time - last_update_time
            last_update_time = current_time
            
            is_attention = check_attention()
            
            # Update duration if attention is true
            if is_attention:
                ATTENTION_DURATION += time_delta
            
            # Track attention switches
            if is_attention and not last_attention_state:
                ATTENTION_SWITCHES += 1
            
            # Record history
            attention_history.append((current_time - session_start_time, is_attention))
            
            # Update UI
            # Attention tracking complete
            
            last_attention_state = is_attention
            time.sleep(0.1)  # Check every 100ms
            
        except Exception as e:
            print(f"Attention tracking error: {e}")
            time.sleep(0.1)

def animal_display_thread():
    """Cycle through animals continuously until all 14 are shown"""
    global current_animal_idx, game_running
    
    DISPLAY_TIME = 10  # seconds per animal
    animals_displayed = 0
    
    while game_running and animals_displayed < len(animal_list):
        try:
            animal = animal_list[current_animal_idx % len(animal_list)]
            show_animal(animal)
            current_animal_idx += 1
            animals_displayed += 1
            time.sleep(DISPLAY_TIME)
            
        except Exception as e:
            print(f"Display thread error: {e}")
            time.sleep(1)
    
    # All 14 animals displayed, end session
    if game_running:
        root.after(500, end_session)

def save_attention_stats():
    """Save attention metrics to Excel"""
    try:
        total_duration = time.time() - session_start_time
        sustained_attention_score = (ATTENTION_DURATION / total_duration * 100) if total_duration > 0 else 0
        
        if os.path.exists(ATTENTION_FILE):
            wb = load_workbook(ATTENTION_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Level 0 Attention"
            
            headers = ["Name", "Attention Duration (s)", "Total Duration (s)", 
                      "Sustained Attention Score (%)", "Attention Frequency", "Date", "Time"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="FF6F00", end_color="FF6F00", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 28
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
        
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        new_row = [
            child_name,
            round(ATTENTION_DURATION, 2),
            round(total_duration, 2),
            round(sustained_attention_score, 2),
            ATTENTION_SWITCHES,
            date_str,
            time_str
        ]
        ws.append(new_row)
        
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(ATTENTION_FILE)
        print(f"✓ Attention stats saved to {ATTENTION_FILE}")
        return True
    except Exception as e:
        print(f"✗ Error saving stats: {e}")
        return False

def show_stats_screen():
    """Display attention statistics"""
    for widget in root.winfo_children():
        widget.pack_forget()
    
    stats_frame = tk.Frame(root, bg="#F4FFDB")
    stats_frame.pack(expand=True, fill="both")
    
    title = tk.Label(
        stats_frame,
        text="Level 0 Attention Analysis",
        font=("Arial", 32, "bold"),
        bg="#F4FFDB",
        fg="#1d3557"
    )
    title.pack(pady=20)
    
    total_duration = time.time() - session_start_time
    sustained_score = (ATTENTION_DURATION / total_duration * 100) if total_duration > 0 else 0
    
    stats_text = f"Focus Time: {ATTENTION_DURATION:.2f}s\n"
    stats_text += f"Total Session: {total_duration:.2f}s\n"
    stats_text += f"Sustained Attention Score: {sustained_score:.2f}%\n"
    stats_text += f"Attention Switches: {ATTENTION_SWITCHES}"
    
    stats_label = tk.Label(
        stats_frame,
        text=stats_text,
        font=("Arial", 20),
        bg="#F4FFDB",
        fg="#1d3557",
        justify="left"
    )
    stats_label.pack(pady=30, padx=20)
    
    continue_btn = tk.Button(
        stats_frame,
        text="Continue to Level 1",
        font=("Arial", 24, "bold"),
        bg="#27ae60",
        fg="white",
        activebackground="#229954",
        activeforeground="white",
        relief="raised",
        padx=30,
        pady=15,
        command=go_to_level1,
        cursor="hand2"
    )
    continue_btn.pack(pady=20)

def go_to_level1():
    """Progress to Level 1"""
    pygame.mixer.music.stop()
    if camera is not None:
        camera.release()
    root.destroy()
    subprocess.run([sys.executable, os.path.join(BASE_DIR, "a_level1_code.py"), child_name])

def exit_app(event=None):
    """Exit application"""
    global game_running, camera
    
    print("Exiting Level 0...")
    game_running = False
    
    pygame.mixer.music.stop()
    if camera is not None:
        camera.release()
    
    root.destroy()
    subprocess.run([sys.executable, os.path.join(BASE_DIR, "main_launcher.py")])

def show_name_entry_screen():
    """Display name entry screen"""
    global child_name
    
    for widget in root.winfo_children():
        widget.destroy()
    
    entry_frame = tk.Frame(root, bg="#F4FFDB")
    entry_frame.pack(expand=True)
    
    title = tk.Label(
        entry_frame,
        text="Level 0: Continuous Learning\n🎓 Watch & Learn! 🎓",
        font=("Arial", 32, "bold"),
        bg="#F4FFDB",
        fg="#1d3557"
    )
    title.pack(pady=30)
    
    name_prompt = tk.Label(
        entry_frame,
        text="Enter Your Name:",
        font=("Arial", 24, "bold"),
        bg="#F4FFDB",
        fg="#1d3557"
    )
    name_prompt.pack(pady=10)
    
    name_entry = tk.Entry(
        entry_frame,
        font=("Arial", 22),
        width=20,
        justify="center",
        bg="white",
        fg="#1d3557",
        relief="solid",
        bd=2
    )
    name_entry.pack(pady=10)
    
    error_label = tk.Label(
        entry_frame,
        text="",
        font=("Arial", 16, "bold"),
        bg="#F4FFDB",
        fg="#e74c3c"
    )
    error_label.pack(pady=5)
    
    def start_clicked():
        global child_name, game_running, session_start_time
        name = name_entry.get().strip()
        if not name:
            error_label.config(text="⚠ Please enter your name!")
            name_entry.config(bg="#ffe6e6")
            root.after(2000, lambda: name_entry.config(bg="white"))
            root.after(2000, lambda: error_label.config(text=""))
        else:
            child_name = name
            print(f"Level 0 started for: {child_name}")
            
            # Initialize camera
            if initialize_camera():
                # Setup main game UI
                for widget in root.winfo_children():
                    widget.destroy()
                
                global image_label, status_label
                
                image_label = tk.Label(root, bg="#F4FFDB")
                image_label.pack(expand=True, fill="both")
                
                status_frame = tk.Frame(root, bg="#F4FFDB")
                status_frame.pack(fill="x", padx=10, pady=10)
                
                status_label = tk.Label(
                    status_frame,
                    text="",
                    font=("Arial", 20, "bold"),
                    bg="#F4FFDB",
                    fg="#1d3557"
                )
                status_label.pack(side="left")
                
                # Control buttons
                controls_frame = tk.Frame(root, bg="#F4FFDB")
                controls_frame.pack(fill="x", padx=10, pady=5)
                
                end_btn = tk.Button(
                    controls_frame,
                    text="End Session",
                    font=("Arial", 18, "bold"),
                    bg="#e63946",
                    fg="white",
                    activebackground="#d62828",
                    command=end_session,
                    cursor="hand2"
                )
                end_btn.pack(side="right", padx=5)
                
                home_btn = tk.Button(
                    controls_frame,
                    text="Home",
                    font=("Arial", 18, "bold"),
                    bg="#3498db",
                    fg="white",
                    activebackground="#2980b9",
                    command=exit_app,
                    cursor="hand2"
                )
                home_btn.pack(side="right", padx=5)
                
                # Start threads
                game_running = True
                session_start_time = time.time()
                
                threading.Thread(target=animal_display_thread, daemon=True).start()
                threading.Thread(target=attention_tracking_thread, daemon=True).start()
                print("✓ Level 0 session started")
            else:
                error_label.config(text="⚠ Camera initialization failed!")
    
    name_entry.bind("<Return>", lambda e: start_clicked())
    root.after(100, lambda: name_entry.focus_force())
    
    start_btn = tk.Button(
        entry_frame,
        text="START",
        font=("Arial", 24, "bold"),
        bg="#27ae60",
        fg="white",
        activebackground="#229954",
        activeforeground="white",
        relief="raised",
        padx=40,
        pady=15,
        command=start_clicked,
        cursor="hand2"
    )
    start_btn.pack(pady=20)

def end_session():
    """End current session and show stats"""
    global game_running
    
    game_running = False
    time.sleep(0.5)  # Wait for threads to finish
    
    # Save stats
    save_attention_stats()
    
    # Show stats screen
    show_stats_screen()

# Main Setup
root = tk.Tk()
root.title("Level 0 - Continuous Learning with Attention Tracking")
root.attributes("-fullscreen", True)
root.bind("<Escape>", exit_app)
root.bind("<Control-q>", exit_app)
root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")
root.configure(bg="#F4FFDB")
root.focus_force()

# Start the app
if child_name:
    # Name provided - initialize and start
    if initialize_camera():
        # Setup UI
        image_label = tk.Label(root, bg="#F4FFDB")
        image_label.pack(expand=True, fill="both")
        
        status_frame = tk.Frame(root, bg="#F4FFDB")
        status_frame.pack(fill="x", padx=10, pady=10)
        
        status_label = tk.Label(status_frame, text="", font=("Arial", 20, "bold"), bg="#F4FFDB", fg="#1d3557")
        status_label.pack(side="left")
        
        # Control buttons
        controls_frame = tk.Frame(root, bg="#F4FFDB")
        controls_frame.pack(fill="x", padx=10, pady=5)
        
        end_btn = tk.Button(controls_frame, text="End Session", font=("Arial", 18, "bold"), bg="#e63946", fg="white", activebackground="#d62828", command=end_session, cursor="hand2")
        end_btn.pack(side="right", padx=5)
        
        home_btn = tk.Button(controls_frame, text="Home", font=("Arial", 18, "bold"), bg="#3498db", fg="white", activebackground="#2980b9", command=exit_app, cursor="hand2")
        home_btn.pack(side="right", padx=5)
        
        # Start session
        game_running = True
        session_start_time = time.time()
        threading.Thread(target=animal_display_thread, daemon=True).start()
        threading.Thread(target=attention_tracking_thread, daemon=True).start()
        print("✓ Level 0 session started")
else:
    # No name - show entry screen
    show_name_entry_screen()

root.mainloop()
