import pygame
import serial
import subprocess
import tkinter as tk
from tkinter import ttk
from PIL import Image, ImageTk
import os
import sys
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

print("=== Level 1: Learning Mode Started ===")

SERIAL_PORT = '/dev/ttyUSB0'
BAUD_RATE = 9600

BASE_DIR = "/home/pi/asd_learning_system"
IMAGE_PATH = os.path.join(BASE_DIR, "animal_images")
SOUND_PATH = os.path.join(BASE_DIR, "animal_sounds")
DEFAULT_IMAGE = os.path.join(IMAGE_PATH, "default_image.jpg")
STATS_FILE = os.path.join(BASE_DIR, "level1_interaction_stats.xlsx")

# Get name from command line (or show entry screen later)
child_name = sys.argv[1] if len(sys.argv) > 1 else ""
if child_name:
    print(f"Level 1 started for: {child_name}")
else:
    print("Level 1: Waiting for name entry...")

# Tracking statistics
animal_tap_count = {}  # {animal_name: tap_count}
animal_time_total = {}  # {animal_name: total_time_spent}
animal_tap_records = []  # list of dicts: {'animal','tap_index','duration','date','time'}
current_animal = None  # Currently displayed animal
animal_start_time = None  # When current animal was shown
current_tap_index = None  # tap index for the currently shown animal

animal_data = {
    "936FA320": "cat",
    "33719E20": "horse",
    "C376B420": "lion",
    "536E4BE4": "parrot",
    "33A8B920": "dog",
    "339FFC30": "sheep",
    "03548F22": "donkey",
    "131CF430": "monkey",
    "236A8222": "duck",
    "135C7A31": "crow",
    "73C28131": "cow",
    "E3257622": "dolphin",
    "23EB9022": "frog",
    "033D9423": "chicken"
}

try:
    ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1)
    print("Serial connected")
except:
    print("Serial error")
    sys.exit()

pygame.init()
pygame.mixer.init()

def exit_app(event=None):
    print("Exiting Level 1...")
    # Finalize tracking for current animal
    global current_animal, animal_start_time
    if current_animal is not None and animal_start_time is not None:
        time_spent = time.time() - animal_start_time
        # record as an individual tap
        now = datetime.now()
        animal_tap_records.append({
            'animal': current_animal,
            'tap_index': current_tap_index,
            'duration': round(time_spent, 2),
            'date': now.strftime("%Y-%m-%d"),
            'time': now.strftime("%H:%M:%S")
        })
        animal_time_total[current_animal] = animal_time_total.get(current_animal, 0) + time_spent
    
    pygame.mixer.music.stop()
    if ser.is_open:
        ser.close()
    root.destroy()
    # Return to launcher
    subprocess.run([sys.executable, os.path.join(BASE_DIR, "main_launcher.py")])

def go_to_level2():
    """Save stats, show them, then progress to Level 2"""
    # Finalize tracking for current animal
    global current_animal, animal_start_time
    if current_animal is not None and animal_start_time is not None:
        time_spent = time.time() - animal_start_time
        now = datetime.now()
        animal_tap_records.append({
            'animal': current_animal,
            'tap_index': current_tap_index,
            'duration': round(time_spent, 2),
            'date': now.strftime("%Y-%m-%d"),
            'time': now.strftime("%H:%M:%S")
        })
        animal_time_total[current_animal] = animal_time_total.get(current_animal, 0) + time_spent
    
    # Save stats to Excel
    save_stats()
    
    # Show stats screen
    show_stats_screen()

def save_stats():
    """Save interaction stats to Excel file"""
    try:
        if os.path.exists(STATS_FILE):
            wb = load_workbook(STATS_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Level 1 Stats"
            
            headers = ["Name", "Animal", "Tap #", "Avg Duration (s)", "Total Duration (s)", "Date", "Time"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
        
        # compute total and average duration per animal from individual tap records
        total_per_animal = {}
        count_per_animal = {}
        for rec in animal_tap_records:
            a = rec['animal']
            total_per_animal[a] = total_per_animal.get(a, 0) + rec['duration']
            count_per_animal[a] = count_per_animal.get(a, 0) + 1

        avg_per_animal = {a: (total_per_animal[a] / count_per_animal[a]) for a in total_per_animal}

        # Add a row for each tap (do NOT combine taps) with avg and total per animal
        for rec in animal_tap_records:
            a = rec['animal']
            new_row = [
                child_name,
                a,
                rec.get('tap_index', ''),
                round(avg_per_animal.get(a, rec['duration']), 2),
                round(total_per_animal.get(a, rec['duration']), 2),
                rec.get('date', ''),
                rec.get('time', '')
            ]
            ws.append(new_row)
            row_num = ws.max_row
            for cell in ws[row_num]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(STATS_FILE)
        print(f"✓ Stats saved to {STATS_FILE}")
        return True
    except Exception as e:
        print(f"✗ Error saving stats: {e}")
        return False

def show_stats_screen():
    """Display interaction statistics before proceeding to Level 2"""
    global root
    
    root.after(1000)  # Brief delay
    
    # Clear previous widgets
    for widget in root.winfo_children():
        widget.pack_forget()
    
    # Stats container
    stats_frame = tk.Frame(root, bg="#F4FFDB")
    stats_frame.pack(expand=True, fill="both")
    
    # Title
    title = tk.Label(
        stats_frame,
        text="Level 1 Interaction Statistics",
        font=("Arial", 32, "bold"),
        bg="#F4FFDB",
        fg="#1d3557"
    )
    title.pack(pady=20)
    
    # Stats info
    total_taps = len(animal_tap_records)
    total_time = sum([rec['duration'] for rec in animal_tap_records])
    unique_animals = len(set([rec['animal'] for rec in animal_tap_records]))

    summary_text = f"Total Animals Explored: {unique_animals}    Total Taps: {total_taps}    Total Time: {total_time:.2f}s"
    summary_label = tk.Label(
        stats_frame,
        text=summary_text,
        font=("Arial", 18, "bold"),
        bg="#F4FFDB",
        fg="#1d3557",
        justify="center"
    )
    summary_label.pack(pady=10)

    # Compute total and average per animal
    total_per_animal = {}
    count_per_animal = {}
    for rec in animal_tap_records:
        a = rec['animal']
        total_per_animal[a] = total_per_animal.get(a, 0) + rec['duration']
        count_per_animal[a] = count_per_animal.get(a, 0) + 1

    avg_per_animal = {a: (total_per_animal[a] / count_per_animal[a]) for a in total_per_animal}

    # Table (Treeview)
    cols = ("Tap #", "Animal", "Avg Duration (s)", "Total Duration (s)", "Date", "Time")
    tree = ttk.Treeview(stats_frame, columns=cols, show='headings', height=10)
    for c in cols:
        tree.heading(c, text=c)
        tree.column(c, anchor='center')
    tree.pack(padx=20, pady=10, expand=True, fill='both')

    # Insert rows for each tap
    for rec in animal_tap_records:
        a = rec['animal']
        tree.insert('', 'end', values=(
            rec.get('tap_index', ''),
            a.upper(),
            f"{avg_per_animal.get(a, rec['duration']):.2f}",
            f"{total_per_animal.get(a, rec['duration']):.2f}",
            rec.get('date', ''),
            rec.get('time', '')
        ))
    
    # Continue button
    continue_btn = tk.Button(
        stats_frame,
        text="Continue to Level 2",
        font=("Arial", 24, "bold"),
        bg="#27ae60",
        fg="white",
        activebackground="#229954",
        activeforeground="white",
        relief="raised",
        padx=30,
        pady=15,
        command=lambda: proceed_to_level2(),
        cursor="hand2"
    )
    continue_btn.pack(pady=20)

def proceed_to_level2():
    """Proceed to Level 2 after stats are shown"""
    print(f"Progressing to Level 2 for {child_name}")
    pygame.mixer.music.stop()
    if ser.is_open:
        ser.close()
    root.destroy()
    subprocess.run([sys.executable, os.path.join(BASE_DIR, "cv_level2_original.py"), child_name])

root = tk.Tk()
root.title("Level 1 - Learning Mode")
root.attributes("-fullscreen", True)
root.bind("<Escape>", exit_app)
root.bind("<Control-q>", exit_app)
root.overrideredirect(True)
root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")
root.configure(bg="#ffffff")
root.focus_force()

def play_sound(animal):
    sound_file = os.path.join(SOUND_PATH, f"{animal}_sound.mp3")
    try:
        pygame.mixer.music.load(sound_file)
        pygame.mixer.music.play()
        pygame.mixer.music.set_volume(1.0)
    except:
        pass

def check_rfid():
    """Poll the serial port for RFID scans and show corresponding animal."""
    try:
        if ser and ser.is_open:
            raw = ser.readline()
            if raw:
                try:
                    s = raw.decode(errors='ignore').strip()
                except:
                    s = str(raw).strip()
                # Cleanup UID string: keep alphanumerics and uppercase
                uid = ''.join(ch for ch in s if ch.isalnum()).upper()
                if uid:
                    # Debug print
                    print(f"RFID read: '{s}' -> UID: {uid}")
                    if uid in animal_data:
                        animal = animal_data[uid]
                        show_animal(animal)
                    else:
                        print(f"Unknown UID: {uid}")
    except Exception as e:
        print(f"RFID read error: {e}")
    finally:
        # Schedule next poll
        try:
            root.after(300, check_rfid)
        except Exception:
            pass

def show_animal(animal):
    """Display animal image/sound and track interaction"""
    global current_animal, animal_start_time
    global current_tap_index

    # If there is an ongoing tap, finalize it (treat each tap separately)
    if current_animal is not None and animal_start_time is not None:
        time_spent = time.time() - animal_start_time
        now = datetime.now()
        animal_tap_records.append({
            'animal': current_animal,
            'tap_index': current_tap_index,
            'duration': round(time_spent, 2),
            'date': now.strftime("%Y-%m-%d"),
            'time': now.strftime("%H:%M:%S")
        })
        animal_time_total[current_animal] = animal_time_total.get(current_animal, 0) + time_spent
        print(f"  {current_animal}: +{time_spent:.2f}s (total: {animal_time_total[current_animal]:.2f}s)")
    
    # Start tracking new animal: increment tap count and set current tap index
    current_animal = animal
    animal_tap_count[animal] = animal_tap_count.get(animal, 0) + 1
    current_tap_index = animal_tap_count[animal]
    animal_start_time = time.time()
    
    print(f"Showing animal: {animal} (tap #{animal_tap_count[animal]})")
    
    try:
        img_path = os.path.join(IMAGE_PATH, f"{animal}_image.jpg")
        img = Image.open(img_path).resize(
            (root.winfo_screenwidth(), root.winfo_screenheight()-100)
        )
        photo = ImageTk.PhotoImage(img)
        image_label.config(image=photo)
        image_label.image = photo
        name_label.config(text=animal.upper())
        play_sound(animal)
    except Exception as e:
        print(f"Error showing animal: {e}")

def show_name_entry_screen():
    """Display name entry screen"""
    global child_name
    
    for widget in root.winfo_children():
        widget.destroy()
    
    entry_frame = tk.Frame(root, bg="#F4FFDB")
    entry_frame.pack(expand=True)
    
    title = tk.Label(
        entry_frame,
        text="Level 1: Learning Mode\n🦁 Say Hello to Animals! 🦁",
        font=("Arial", 32, "bold"),
        bg="#F4FFDB",
        fg="#1d3557"
    )
    title.pack(pady=30)
    
    name_prompt = tk.Label(
        entry_frame,
        text="Enter Your Name:",
        font=("Arial", 24, "bold"),
        bg="#F4FFDB",
        fg="#1d3557"
    )
    name_prompt.pack(pady=10)
    
    name_entry = tk.Entry(
        entry_frame,
        font=("Arial", 22),
        width=20,
        justify="center",
        bg="white",
        fg="#1d3557",
        relief="solid",
        bd=2
    )
    name_entry.pack(pady=10)
    
    error_label = tk.Label(
        entry_frame,
        text="",
        font=("Arial", 16, "bold"),
        bg="#F4FFDB",
        fg="#e74c3c"
    )
    error_label.pack(pady=5)
    
    def start_clicked():
        global child_name
        name = name_entry.get().strip()
        if not name:
            error_label.config(text="⚠ Please enter your name!")
            name_entry.config(bg="#ffe6e6")
            root.after(2000, lambda: name_entry.config(bg="white"))
            root.after(2000, lambda: error_label.config(text=""))
        else:
            child_name = name
            print(f"Level 1 started for: {child_name}")
            start_learning()
    
    name_entry.bind("<Return>", lambda e: start_clicked())
    root.after(100, lambda: name_entry.focus_force())
    
    start_btn = tk.Button(
        entry_frame,
        text="START",
        font=("Arial", 24, "bold"),
        bg="#27ae60",
        fg="white",
        activebackground="#229954",
        activeforeground="white",
        relief="raised",
        padx=40,
        pady=15,
        command=start_clicked,
        cursor="hand2"
    )
    start_btn.pack(pady=20)

def start_learning():
    """Start the main learning game"""
    for widget in root.winfo_children():
        widget.destroy()
    
    # Default image
    default_img = Image.open(DEFAULT_IMAGE).resize(
        (root.winfo_screenwidth(), root.winfo_screenheight()-100)
    )
    default_photo_obj = ImageTk.PhotoImage(default_img)
    
    global image_label, name_label
    image_label = tk.Label(root, image=default_photo_obj, bg="#F4FFDB")
    image_label.pack()
    image_label.image = default_photo_obj
    
    name_label = tk.Label(root, text="", font=("Arial", 40, "bold"), bg="#ffffff")
    name_label.pack()
    
    # Control buttons frame (top-right corner)
    controls_frame = tk.Frame(root, bg="#FFFFFF")
    controls_frame.place(relx=0.98, rely=0.02, anchor="ne")
    
    # Next Level button
    next_level_btn = tk.Button(
        controls_frame,
        text="Next Level",
        font=("Arial", 18, "bold"),
        bg="#27ae60",
        fg="white",
        activebackground="#229954",
        activeforeground="white",
        relief="raised",
        command=go_to_level2,
        cursor="hand2"
    )
    next_level_btn.pack(side="right", padx=5)
    
    # Home button
    home_btn = tk.Button(
        controls_frame,
        text="Home",
        font=("Arial", 18, "bold"),
        bg="#3498db",
        fg="white",
        activebackground="#2980b9",
        activeforeground="white",
        relief="raised",
        command=exit_app,
        cursor="hand2"
    )
    home_btn.pack(side="right", padx=5)
    
    # Start RFID checking
    check_rfid()

# Start the app
if child_name:
    # Name provided from command line - start learning directly
    start_learning()
else:
    # No name provided - show entry screen
    show_name_entry_screen()
root.mainloop()
