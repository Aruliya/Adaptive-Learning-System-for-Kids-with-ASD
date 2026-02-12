import pygame
import serial
import random
import subprocess
import tkinter as tk
from PIL import Image, ImageTk, ImageSequence
import os
import sys
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import cv2
import mediapipe as mp
import time

print("=== Level 3: Audio Matching Mode Started ===")

SERIAL_PORT = '/dev/ttyUSB0'
BAUD_RATE = 9600

BASE_DIR = "/home/pi/asd_learning_system"
IMAGE_PATH = os.path.join(BASE_DIR, "animal_images")
SOUND_PATH = os.path.join(BASE_DIR, "animal_sounds")
GIF_PATH = os.path.join(BASE_DIR, "feedback_gifs")
RESULTS_FILE = os.path.join(BASE_DIR, "level3_results.xlsx")
ATTENTION_FILE = os.path.join(BASE_DIR, "attention_tracker_level3.xlsx")

TOTAL_QUESTIONS = 14  
MAX_RETRIES = 3
LOOKDOWN_WINDOW = 15  # 15 seconds lookdown window
gif_running = False
current_gif_frames = []
accepting_input = False

# Get name from command line or show entry screen
child_name = sys.argv[1] if len(sys.argv) > 1 else ""

# Attention tracking - GAME-WIDE
total_attention_duration = 0  # Total time face was detected during all questions
total_game_duration = 0       # Total time for all questions combined
game_start_time = 0
is_in_lookdown = False
lookdown_start_time = 0
face_detected = False

# MediaPipe Face Detection (setup deferred until camera init)
mp_face_detection = mp.solutions.face_detection
face_detection = None
# Detection tuning
DETECTION_CONFIDENCE = 0.2
# seconds to treat a recent detection as still present (smoothing)
DETECTION_DECAY = 0.35
# timestamp of last positive detection
last_face_timestamp = 0

# Camera
camera = None
camera_running = False
last_attention_check = 0

# Audio stimuli timing for Level 3
stimulus_start_time = 0  # When audio starts playing (stimuli shown)
stimulus_end_time = 0    # When audio ends or next stimulus begins

animal_data = {
    "936FA320": "cat",
    "33719E20": "horse",
    "C376B420": "lion",
    "536E4BE4": "parrot",
    "33A8B920": "dog",
    "339FFC30": "sheep",
    "03548F22": "donkey",
    "131CF430": "monkey",
    "236A8222": "duck",
    "135C7A31": "crow",
    "73C28131": "cow",
    "E3257622": "dolphin",
    "23EB9022": "frog",
    "033D9423": "chicken"
}

animal_to_uid = {v: k for k, v in animal_data.items()}
animal_list = list(animal_to_uid.keys())
shuffled_animals = []  # Will store shuffled list of animals for the game

try:
    ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1)
except:
    print("Serial error")
    sys.exit()

pygame.init()
pygame.mixer.init()

def go_home():
    """Return to launcher"""
    print("Returning to home...")
    pygame.mixer.music.stop()
    if ser.is_open:
        ser.close()
    root.destroy()
    subprocess.run([sys.executable, os.path.join(BASE_DIR, "main_launcher.py")])

def exit_app(event=None):
    print("Exiting application safely...")
    global accepting_input
    accepting_input = False
    
    pygame.mixer.music.stop()
    
    if ser.is_open:
        ser.close()
    root.quit()
    root.destroy()

root = tk.Tk()
root.attributes("-fullscreen", True)
root.bind("<Escape>", exit_app)
root.bind("<Control-q>", exit_app)
# COMMENTED OUT - This prevents keyboard input!
# root.overrideredirect(True)  
root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")
root.configure(bg="#F4FFDB")
root.focus_force()

current_animal = None
question_count = 0
retry_count = 0
score = 0

# Main container
main_frame = tk.Frame(root, bg="#F4FFDB")
main_frame.pack(expand=True, fill="both")

image_label = tk.Label(main_frame, bg="#F4FFDB")
image_label.pack(expand=True)

status_label = tk.Label(main_frame, font=("Arial", 28), bg="#F4FFDB")
status_label.pack()

# Attention indicator label
attention_label = tk.Label(main_frame, font=("Arial", 16), bg="#F4FFDB", fg="#006064")
attention_label.pack()

# ---------- FACE DETECTION FUNCTIONS ----------

def initialize_camera():
    """Initialize the camera for face detection"""
    global camera, camera_running
    global face_detection, last_face_timestamp
    try:
        # Try multiple indexes (some Pis need 1 or 2)
        for idx in range(0, 4):
            cam = cv2.VideoCapture(idx)
            if cam is not None and cam.isOpened():
                camera = cam
                break

        if camera is None or not camera.isOpened():
            print("✗ Failed to open any camera index")
            return False

        # Set a reasonable resolution for face detection
        camera.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        camera.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

        # Initialize MediaPipe face detector with tuned confidence
        try:
            if face_detection is not None:
                face_detection.close()
        except Exception:
            pass

        face_detection = mp_face_detection.FaceDetection(
            model_selection=0,
            min_detection_confidence=DETECTION_CONFIDENCE
        )

        last_face_timestamp = 0
        camera_running = True
        print("✓ Camera initialized for face detection (index set and MediaPipe ready)")
        return True
    except Exception as e:
        print(f"✗ Camera error: {e}")
        return False

def detect_face_in_frame():
    """
    Process a single camera frame for face detection
    Returns: True if face detected, False otherwise
    """
    global face_detected, camera_running
    global face_detection, last_face_timestamp

    if not camera_running or camera is None:
        return False
   
    ret, frame = camera.read()
    if not ret:
        return False
   
    try:
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        # Let MediaPipe work on a read-only buffer for performance
        rgb_frame.flags.writeable = False

        if face_detection is None:
            # No detector available (camera init failed to create it)
            face_detected = False
            return False

        results = face_detection.process(rgb_frame)

        # Restore writeable flag (not strictly necessary here)
        rgb_frame.flags.writeable = True

        if results and getattr(results, 'detections', None):
            # Update timestamp of last positive detection
            last_face_timestamp = time.time()
            face_detected = True
            return True
        else:
            face_detected = False
            return False
    except Exception as e:
        print(f"Face detection processing error: {e}")
        return False

def attention_tracking_loop():
    """
    Continuous loop for face detection and attention tracking
    Runs in separate thread
    For Level 3 (audio mode): Track attention during audio playback until next stimulus
    """
    global total_attention_duration, face_detected, last_attention_check
    global last_face_timestamp, stimulus_start_time, stimulus_end_time

    last_time = time.time()

    while camera_running:
        try:
            current_time = time.time()
            time_delta = current_time - last_time
            last_time = current_time

            # Run a detection pass (updates last_face_timestamp)
            _ = detect_face_in_frame()

            # Consider face present if detected recently (smoothing)
            has_face = (time.time() - last_face_timestamp) <= DETECTION_DECAY

            # Update attention duration - include time during lookdown if face detected
            # For audio mode: count actual attention time whenever face is visible, including lookdown window
            if stimulus_start_time > 0 and has_face:
                total_attention_duration += time_delta

            # Update UI status
            if is_in_lookdown:
                status_text = "🎧 Listening..."
            elif has_face:
                status_text = "😊 Face Detected ✓"
            else:
                status_text = "😶 No Face Detected"

            # Update label (thread-safe)
            try:
                attention_label.config(text=status_text)
            except:
                pass

            # Update every ~50ms (20 FPS)
            time.sleep(0.05)

        except Exception as e:
            print(f"Face detection error: {e}")
            time.sleep(0.1)

def start_game_tracking():
    """Start tracking when game begins"""
    global game_start_time, total_attention_duration, total_game_duration
   
    game_start_time = time.time()
    total_attention_duration = 0
    total_game_duration = 0
   
    print("⏱️ Game-wide attention tracking started")

def trigger_lookdown_window():
    """
    Start the 15-second lookdown window
    During this time, not having a face detected doesn't count against attention
    """
    global is_in_lookdown, lookdown_start_time
   
    is_in_lookdown = True
    lookdown_start_time = time.time()
    print(f"🎧 Lookdown window started ({LOOKDOWN_WINDOW} seconds)")
   
    # Schedule end of lookdown window
    root.after(LOOKDOWN_WINDOW * 1000, end_lookdown_window)

def end_lookdown_window():
    """End the lookdown window"""
    global is_in_lookdown
   
    is_in_lookdown = False
    print("🎧 Lookdown window ended")

def stop_game_tracking():
    """
    Stop tracking and calculate final metrics
    Returns: (attention_duration, total_duration, attention_score)
    """
    global total_attention_duration, total_game_duration, game_start_time
   
    if game_start_time == 0:
        return 0, 0, 0
   
    # Calculate total game duration
    total_game_duration = time.time() - game_start_time
   
    # Calculate attention score (percentage)
    if total_game_duration > 0:
        attention_score = (total_attention_duration / total_game_duration) * 100
    else:
        attention_score = 0
   
    attention_score = round(attention_score, 2)
    attention_duration_rounded = round(total_attention_duration, 2)
    total_duration_rounded = round(total_game_duration, 2)
   
    print(f"\n{'='*60}")
    print(f"📊 FINAL ATTENTION METRICS (LEVEL 3)")
    print(f"{'='*60}")
    print(f"   Total Attention Duration: {attention_duration_rounded}s")
    print(f"   Total Game Duration: {total_duration_rounded}s")
    print(f"   Attention Score: {attention_score}%")
    print(f"{'='*60}\n")
   
    return attention_duration_rounded, total_duration_rounded, attention_score

# ---------- UTILITIES ----------

def save_results_to_excel():
    """Save the game results to Excel file"""
    try:
        if os.path.exists(RESULTS_FILE):
            wb = load_workbook(RESULTS_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Level 3 Results"
            
            headers = ["Name", "Score", "Total Questions", "Percentage", "Date", "Time"]
            ws.append(headers)
            
            header_fill = PatternFill(start_color="00ACC1", end_color="00ACC1", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
        
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        percentage = round((score / TOTAL_QUESTIONS) * 100, 1)
        
        new_row = [child_name, score, TOTAL_QUESTIONS, f"{percentage}%", date_str, time_str]
        ws.append(new_row)
        
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        wb.save(RESULTS_FILE)
        print(f"Results saved: {child_name} - {score}/{TOTAL_QUESTIONS} ({percentage}%) - {date_str} {time_str}")
        return True
    except Exception as e:
        print(f"Error saving results: {e}")
        return False

def save_attention_results(attention_duration, total_duration, attention_score):
    """Save attention metrics to separate attention tracker file"""
    try:
        if os.path.exists(ATTENTION_FILE):
            wb = load_workbook(ATTENTION_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Level 3 Attention Tracking"
           
            headers = ["Name", "Attention Duration (s)", "Total Duration (s)", "Attention Score (%)", "Date", "Time"]
            ws.append(headers)
           
            header_fill = PatternFill(start_color="00ACC1", end_color="00ACC1", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
           
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
           
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 22
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 22
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
       
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
       
        new_row = [
            child_name,
            attention_duration,
            total_duration,
            attention_score,
            date_str,
            time_str
        ]
        ws.append(new_row)
       
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
       
        wb.save(ATTENTION_FILE)
        print(f"✓ Attention data saved to {ATTENTION_FILE}")
        print(f"  Entry: {child_name} - {attention_duration}s / {total_duration}s ({attention_score}%)")
       
        return True
    except Exception as e:
        print(f"✗ Error saving attention data: {e}")
        return False

def play_audio_non_blocking(file):
    """Play audio in a separate thread"""
    def play():
        try:
            pygame.mixer.music.load(file)
            pygame.mixer.music.play()
        except:
            pass
    threading.Thread(target=play, daemon=True).start()

def play_audio_blocking(file):
    try:
        pygame.mixer.music.load(file)
        pygame.mixer.music.play()
        while pygame.mixer.music.get_busy():
            root.update()
    except:
        pass

def play_animal_sound():
    """Play the current animal sound once"""
    if current_animal:
        try:
            sound_file = os.path.join(SOUND_PATH, f"{current_animal}_sound.mp3")
            pygame.mixer.music.load(sound_file)
            pygame.mixer.music.play()  # Play once, no loop
            pygame.mixer.music.set_volume(1.0)
            print(f"Playing {current_animal} sound...")
        except Exception as e:
            print(f"Error playing sound: {e}")

def show_listening_screen():
    """Show a visual indicator that audio is playing with replay button"""
    try:
        # Clear any existing image
        image_label.config(image="")
        image_label.image = None
        
        listen_frame = tk.Frame(main_frame, bg="#F4FFDB")
        listen_frame.pack(expand=True)
        listen_frame.place(relx=0.5, rely=0.5, anchor="center")
        
        guess_audio = tk.Label(
            listen_frame,
            text="Guess the animal!",
            font=("Arial", 44, "bold"),
            bg="#F4FFDB",
            fg="#2c3e50"
        )
        guess_audio.pack(pady=20)
        
        # REPLAY BUTTON
        replay_btn = tk.Button(
            listen_frame,
            text="REPLAY SOUND",
            font=("Arial", 28, "bold"),
            bg="#00acc1",
            fg="white",
            activebackground="#00838f",
            activeforeground="white",
            relief="raised",
            bd=5,
            padx=40,
            pady=15,
            command=play_animal_sound,
            cursor="hand2"
        )
        replay_btn.pack(pady=20)
        
        # Store reference to prevent garbage collection
        image_label.listen_frame = listen_frame
        
    except Exception as e:
        print(f"Error showing listening screen: {e}")

def clear_listening_screen():
    """Clear the listening screen widgets"""
    try:
        if hasattr(image_label, 'listen_frame'):
            image_label.listen_frame.destroy()
            delattr(image_label, 'listen_frame')
    except:
        pass

def show_animal_with_name():
    """Show the animal image with its name"""
    try:
        clear_listening_screen()
        
        img = Image.open(os.path.join(IMAGE_PATH, f"{current_animal}_image.jpg"))
        img = img.resize((root.winfo_screenwidth(), root.winfo_screenheight() - 150))
        photo = ImageTk.PhotoImage(img)
        image_label.config(image=photo)
        image_label.image = photo
        
        # Update status to show animal name
        status_label.config(
            text=f"✓ {current_animal.upper()}!",
            font=("Arial", 48, "bold"),
            fg="#00695c"
        )
    except Exception as e:
        print(f"Error showing animal: {e}")

def show_gif_with_audio(gif_name, audio_file=None, duration=2000, loop=False):
    """Show GIF and play audio simultaneously"""
    global gif_running, current_gif_frames
    
    clear_listening_screen()
    
    gif_running = False
    root.update()
    gif_running = True

    try:
        gif_path = os.path.join(GIF_PATH, gif_name)
        gif_image = Image.open(gif_path)
        current_gif_frames = [
            ImageTk.PhotoImage(frame.resize(
                (root.winfo_screenwidth(), root.winfo_screenheight())))
            for frame in ImageSequence.Iterator(gif_image)
        ]

        if audio_file:
            play_audio_non_blocking(audio_file)

        def animate(idx=0):
            if not gif_running:
                return
            
            if not loop and idx >= len(current_gif_frames):
                if current_gif_frames:
                    image_label.config(image=current_gif_frames[-1])
                return
                
            image_label.config(image=current_gif_frames[idx % len(current_gif_frames)])
            root.after(100, animate, idx + 1)

        animate()
        
        if not loop:
            root.after(duration, stop_gif)
    except Exception as e:
        print(f"Error showing GIF: {e}")

def stop_gif():
    global gif_running
    gif_running = False

# ---------- START SCREEN ----------

def show_start_screen():
    """Display welcome screen with GIF, name entry, and start button"""
    global child_name, accepting_input
    
    accepting_input = False
    
    for widget in main_frame.winfo_children():
        widget.pack_forget()
    
    start_container = tk.Frame(main_frame, bg="#F4FFDB")
    start_container.pack(expand=True)
    
    gif_label = tk.Label(start_container, bg="#F4FFDB")
    gif_label.pack(pady=10)
    
    try:
        gif_path = os.path.join(GIF_PATH, "default_level3.gif")
        gif_image = Image.open(gif_path)
        gif_frames = [
            ImageTk.PhotoImage(frame.resize((350, 250)))
            for frame in ImageSequence.Iterator(gif_image)
        ]
        
        frame_count = [0]
        
        def animate_start_gif():
            if frame_count[0] < len(gif_frames) and gif_label.winfo_exists():
                gif_label.config(image=gif_frames[frame_count[0]])
                gif_label.image = gif_frames[frame_count[0]]
                frame_count[0] += 1
                root.after(100, animate_start_gif)
            elif gif_label.winfo_exists():
                gif_label.config(image=gif_frames[-1])
                gif_label.image = gif_frames[-1]
        
        animate_start_gif()
    except Exception as e:
        print(f"Error loading start GIF: {e}")
    
    name_frame = tk.Frame(start_container, bg="#F4FFDB")
    name_frame.pack(pady=15)
    
    title_label = tk.Label(
        name_frame,
        text="LEVEL 3: Audio Matching",
        font=("Arial", 28, "bold"),
        bg="#F4FFDB",
        fg="#006064"
    )
    title_label.pack(pady=5)
    
    name_label = tk.Label(
        name_frame, 
        text="Enter Your Name:", 
        font=("Arial", 24, "bold"),
        bg="#F4FFDB",
        fg="#00838f"
    )
    name_label.pack(pady=5)
    
    name_entry = tk.Entry(
        name_frame,
        font=("Arial", 22),
        width=20,
        justify="center",
        bg="white",
        fg="#006064",
        relief="solid",
        bd=2,
        insertbackground="#006064"
    )
    name_entry.pack(pady=8)
    
    error_label = tk.Label(
        name_frame,
        text="",
        font=("Arial", 18, "bold"),
        bg="#F4FFDB",
        fg="#e74c3c"
    )
    error_label.pack(pady=5)
    
    def start_game_clicked():
        global child_name, game_start_time, total_attention_duration, total_game_duration
        child_name = name_entry.get().strip()
        if not child_name:
            error_label.config(text="⚠ Please enter your name!")
            name_entry.focus_force()
            name_entry.config(bg="#ffe6e6")
            root.after(2000, lambda: name_entry.config(bg="white"))
            root.after(2000, lambda: error_label.config(text=""))
        else:
            error_label.config(text="")
            print(f"Starting Level 3 for: {child_name}")
            
            # START TIMING: Right after name entry validation
            game_start_time = time.time()
            total_attention_duration = 0
            total_game_duration = 0
            print("⏱️ Game-wide attention tracking started")
            
            start_container.destroy()
            
            # Initialize camera before starting game
            if initialize_camera():
                start_game()
            else:
                print("⚠️ Warning: Camera not available, face detection disabled for Level 3")
                start_game()
    
    name_entry.bind("<Return>", lambda e: start_game_clicked())
    name_entry.bind("<KeyPress>", lambda e: error_label.config(text=""))
    name_entry.bind("<Button-1>", lambda e: name_entry.focus_force())
    
    name_entry.focus_set()
    name_entry.focus_force()
    name_entry.icursor(0)
    root.update()
    root.after(100, lambda: name_entry.focus_force())
    root.after(500, lambda: name_entry.focus_force())
    
    start_btn = tk.Button(
        name_frame,
        text="START",
        font=("Arial", 24, "bold"),
        bg="#00acc1",
        fg="white",
        activebackground="#00838f",
        activeforeground="white",
        relief="raised",
        bd=5,
        padx=30,
        pady=10,
        command=start_game_clicked,
        cursor="hand2"
    )
    start_btn.pack(pady=12)

# ---------- GAME FLOW ----------

def start_game():
    global accepting_input, shuffled_animals, camera_running
    
    # Create shuffled list of all animals for unique questions
    shuffled_animals = animal_list.copy()
    random.shuffle(shuffled_animals)
    
    image_label.pack(expand=True)
    status_label.pack()
    attention_label.pack()
    
    accepting_input = False
    
    # Initialize camera before starting game
    if initialize_camera():
        # Start face detection thread
        tracking_thread = threading.Thread(target=attention_tracking_loop, daemon=True)
        tracking_thread.start()
        print("✓ Face detection thread started for Level 3")
    else:
        print("⚠️ Warning: Camera not available, face detection disabled for Level 3")
    
    show_new_animal()

def show_new_animal():
    global current_animal, retry_count, question_count, accepting_input
    global stimulus_start_time, stimulus_end_time

    if question_count >= TOTAL_QUESTIONS:
        show_final_score()
        return

    stop_gif()
    clear_listening_screen()
    
    retry_count = 0
    
    # Get next unique animal from shuffled list
    current_animal = shuffled_animals[question_count]
    question_count += 1

    print(f"Question {question_count}: {current_animal}")

    try:
        # Show listening screen
        show_listening_screen()
        status_label.config(
            text=f"Question {question_count}/{TOTAL_QUESTIONS}",
            font=("Arial", 28),
            fg="#006064"
        )
        
        # Reset stimulus timing
        stimulus_start_time = 0
        stimulus_end_time = 0
        
        # Start 15-second lookdown window for this question
        trigger_lookdown_window()
        
        # Play animal sound once and track stimulus duration
        root.after(500, lambda: play_animal_sound_and_track())
        
        # Enable input after sound starts
        root.after(2500, lambda: setattr(sys.modules[__name__], 'accepting_input', True))
        
    except Exception as e:
        print(f"Error in show_new_animal: {e}")

def play_animal_sound_and_track():
    """Play animal sound and set stimulus timing for attention calculation"""
    global stimulus_start_time
    stimulus_start_time = time.time()
    play_animal_sound()
    print(f"🔊 Audio stimulus started for {current_animal} at {stimulus_start_time}")

def check_rfid():
    global retry_count, score, accepting_input, stimulus_end_time

    if ser.in_waiting:
        uid = ser.readline().decode(errors="ignore").strip().upper()
        
        if not uid or not accepting_input or current_animal is None:
            root.after(100, check_rfid)
            return
            
        print(f"RFID scanned: {uid}")
        
        # Mark stimulus end time (transition to next stimulus)
        stimulus_end_time = time.time()
        accepting_input = False

        if uid == animal_to_uid[current_animal]:
            score += 1
            
            # Show animal image with name
            show_animal_with_name()
            root.after(2000, lambda: show_gif_with_audio(
                "goodJob.gif",
                os.path.join(SOUND_PATH, "correct_sound.mp3"),
                duration=2000
            ))
            root.after(4500, show_new_animal)

        else:
            retry_count += 1
            if retry_count < MAX_RETRIES:
                show_gif_with_audio(
                    "tryAgain.gif",
                    os.path.join(SOUND_PATH, "incorrect_sound.mp3"),
                    duration=1500
                )
                
                def retry_audio():
                    global stimulus_start_time
                    show_listening_screen()
                    # Play the sound again after error - restart stimulus timer
                    root.after(500, lambda: play_animal_sound_and_track())
                    root.after(2500, lambda: setattr(sys.modules[__name__], 'accepting_input', True))
                
                root.after(1600, retry_audio)
            else:
                show_gif_with_audio(
                    "uhOh.gif",
                    os.path.join(SOUND_PATH, "oops_sound.mp3"),
                    duration=2000
                )
                root.after(2500, show_new_animal)
    
    root.after(100, check_rfid)

def show_final_score():
    """Display final score with attention summary"""
    global gif_running, current_gif_frames, accepting_input, camera_running
    
    accepting_input = False
    camera_running = False  # Stop camera
    
    pygame.mixer.music.stop()
    stop_gif()
    clear_listening_screen()
    
    # Stop tracking and get final metrics
    attention_duration, total_duration, attention_score = stop_game_tracking()
    
    # Save both files
    save_results_to_excel()
    save_attention_results(attention_duration, total_duration, attention_score)
    
    percentage = (score / TOTAL_QUESTIONS) * 100
    
    status_label.pack_forget()
    attention_label.pack_forget()
    image_label.pack_forget()
    
    final_container = tk.Frame(main_frame, bg="#F4FFDB")
    final_container.pack(expand=True)
    
    final_gif_label = tk.Label(final_container, bg="#F4FFDB")
    final_gif_label.pack(pady=20)
    
    # Different message if launched from launcher vs standalone
    if child_name and len(sys.argv) > 1:
        score_text = f"Congratulations, {child_name}! \n\nYou completed all 3 levels!\n\nLevel 3 Score: {score}/{TOTAL_QUESTIONS} ({percentage:.1f}%)\n\n😊 Attention Score: {attention_score}%"
    else:
        score_text = f"Awesome, {child_name}!\n\nYou matched {score} out of {TOTAL_QUESTIONS} sounds correctly!\n\nScore: {percentage:.1f}%\n\n😊 Attention Score: {attention_score}%"
    
    score_label = tk.Label(
        final_container,
        text=score_text,
        font=("Arial", 28, "bold"),
        bg="#F4FFDB",
        fg="#006064",
        justify="center"
    )
    score_label.pack(pady=30)
    
    try:
        gif_path = os.path.join(GIF_PATH, "finalScore.gif")
        gif_image = Image.open(gif_path)
        final_frames = [
            ImageTk.PhotoImage(frame.resize((600, 400)))
            for frame in ImageSequence.Iterator(gif_image)
        ]
        
        frame_count = [0]
        
        def animate_final_gif():
            if frame_count[0] < len(final_frames) and final_gif_label.winfo_exists():
                final_gif_label.config(image=final_frames[frame_count[0]])
                final_gif_label.image = final_frames[frame_count[0]]
                frame_count[0] += 1
                root.after(100, animate_final_gif)
            elif final_gif_label.winfo_exists():
                final_gif_label.config(image=final_frames[-1])
                final_gif_label.image = final_frames[-1]
        
        animate_final_gif()
    except Exception as e:
        print(f"Error loading final GIF: {e}")
    
    buttons_frame = tk.Frame(final_container, bg="#F4FFDB")
    buttons_frame.pack(pady=20)
    
    # Show Home button only if launched from launcher
    if len(sys.argv) > 1:
        home_btn = tk.Button(
            buttons_frame,
            text="Back to Home",
            font=("Arial", 28, "bold"),
            bg="#3498db",
            fg="white",
            activebackground="#2980b9",
            activeforeground="white",
            relief="raised",
            bd=5,
            padx=40,
            pady=15,
            command=go_home,
            cursor="hand2"
        )
        home_btn.pack(side="left", padx=10)
    
    exit_btn = tk.Button(
        buttons_frame,
        text="EXIT",
        font=("Arial", 28, "bold"),
        bg="#e63946",
        fg="white",
        activebackground="#d62828",
        activeforeground="white",
        relief="raised",
        bd=5,
        padx=40,
        pady=15,
        command=exit_app,
        cursor="hand2"
    )
    exit_btn.pack(side="left", padx=10)

# ---------- START ----------
if child_name:
    # Name provided from launcher - skip name entry
    print(f"Level 3 started for: {child_name}")
    start_game()
else:
    # No name provided - show name entry
    show_start_screen()
    
check_rfid()
root.mainloop()
