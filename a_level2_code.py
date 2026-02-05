import pygame
import serial
import random
import subprocess
import tkinter as tk
from PIL import Image, ImageTk, ImageSequence
import os
import sys
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

print("=== Level 2: Image Identification Mode Started ===")

SERIAL_PORT = '/dev/ttyUSB0'
BAUD_RATE = 9600

BASE_DIR = "/home/pi/asd_learning_system"
IMAGE_PATH = os.path.join(BASE_DIR, "animal_images")
SOUND_PATH = os.path.join(BASE_DIR, "animal_sounds")
GIF_PATH = os.path.join(BASE_DIR, "feedback_gifs")
RESULTS_FILE = os.path.join(BASE_DIR, "level2_results.xlsx")

TOTAL_QUESTIONS = 14
MAX_RETRIES = 3
gif_running = False
current_gif_frames = []
accepting_input = False

# Get name from command line or show entry screen
child_name = sys.argv[1] if len(sys.argv) > 1 else ""  # Flag to control RFID scanning

animal_data = {
    "936FA320": "cat",
    "33719E20": "horse",
    "6371A320": "elephant",
    "C376B420": "lion",
    "536E4BE4": "parrot",
    "33A8B920": "dog",
    "339FFC30": "sheep",
    "03548F22": "donkey",
    "131CF430": "monkey",
    "236A8222": "duck",
    "135C7A31": "crow",
    "73C28131": "cow",
    "E3257622": "dolphin",
    "23EB9022": "frog",
    "033D9423": "chicken"
}

animal_to_uid = {v: k for k, v in animal_data.items()}
animal_list = list(animal_to_uid.keys())
shuffled_animals = []  # Will store shuffled list of animals for the game

try:
    ser = serial.Serial(SERIAL_PORT, BAUD_RATE, timeout=1)
except:
    print("Serial error")
    sys.exit()

pygame.init()
pygame.mixer.init()

def exit_app(event=None):
    print("Exiting application safely...")
    global accepting_input
    accepting_input = False
    pygame.mixer.music.stop()
    if ser.is_open:
        ser.close()
    root.quit()
    root.destroy()

root = tk.Tk()
root.attributes("-fullscreen", True)
root.bind("<Escape>", exit_app)
root.bind("<Control-q>", exit_app)
root.overrideredirect(True)
root.geometry(f"{root.winfo_screenwidth()}x{root.winfo_screenheight()}+0+0")
root.configure(bg="#F4FFDB")
root.focus_force()

current_animal = None
question_count = 0
retry_count = 0
score = 0

# Main container
main_frame = tk.Frame(root, bg="#F4FFDB")
main_frame.pack(expand=True, fill="both")

image_label = tk.Label(main_frame, bg="#F4FFDB")
image_label.pack(expand=True)

status_label = tk.Label(main_frame, font=("Arial", 28), bg="#F4FFDB")
status_label.pack()

# ---------- UTILITIES ----------

def save_results_to_excel():
    """Save the game results to Excel file"""
    try:
        # Check if file exists
        if os.path.exists(RESULTS_FILE):
            wb = load_workbook(RESULTS_FILE)
            ws = wb.active
        else:
            # Create new workbook with headers
            wb = Workbook()
            ws = wb.active
            ws.title = "Level 2 Results"
            
            # Add headers with styling
            headers = ["Name", "Score", "Total Questions", "Percentage", "Date", "Time"]
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="457B9D", end_color="457B9D", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
        
        # Get current date and time
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        # Calculate percentage
        percentage = round((score / TOTAL_QUESTIONS) * 100, 1)
        
        # Add new row with results
        new_row = [
            child_name,
            score,
            TOTAL_QUESTIONS,
            f"{percentage}%",
            date_str,
            time_str
        ]
        ws.append(new_row)
        
        # Style the new row
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Save the file
        wb.save(RESULTS_FILE)
        print(f"Results saved successfully to {RESULTS_FILE}")
        print(f"Entry: {child_name} - {score}/{TOTAL_QUESTIONS} ({percentage}%) - {date_str} {time_str}")
        
        return True
    except Exception as e:
        print(f"Error saving results to Excel: {e}")
        return False

def play_audio_non_blocking(file):
    """Play audio in a separate thread"""
    def play():
        try:
            pygame.mixer.music.load(file)
            pygame.mixer.music.play()
        except:
            pass
    threading.Thread(target=play, daemon=True).start()

def play_audio_blocking(file):
    try:
        pygame.mixer.music.load(file)
        pygame.mixer.music.play()
        while pygame.mixer.music.get_busy():
            root.update()
    except:
        pass

def show_current_animal():
    try:
        img = Image.open(os.path.join(IMAGE_PATH, f"{current_animal}_image.jpg"))
        img = img.resize((root.winfo_screenwidth(), root.winfo_screenheight()))
        photo = ImageTk.PhotoImage(img)
        image_label.config(image=photo)
        image_label.image = photo
    except:
        pass

def show_gif_with_audio(gif_name, audio_file=None, duration=2000, loop=False):
    """Show GIF and play audio simultaneously"""
    global gif_running, current_gif_frames
    gif_running = False
    root.update()
    gif_running = True

    try:
        # Load and resize GIF frames
        gif_path = os.path.join(GIF_PATH, gif_name)
        gif_image = Image.open(gif_path)
        current_gif_frames = [
            ImageTk.PhotoImage(frame.resize(
                (root.winfo_screenwidth(), root.winfo_screenheight())))
            for frame in ImageSequence.Iterator(gif_image)
        ]

        # Play audio if provided (non-blocking)
        if audio_file:
            play_audio_non_blocking(audio_file)

        def animate(idx=0):
            if not gif_running:
                return
            
            if not loop and idx >= len(current_gif_frames):
                # GIF played once, show last frame
                if current_gif_frames:
                    image_label.config(image=current_gif_frames[-1])
                return
                
            image_label.config(image=current_gif_frames[idx % len(current_gif_frames)])
            root.after(100, animate, idx + 1)

        animate()
        
        if not loop:
            root.after(duration, stop_gif)
    except Exception as e:
        print(f"Error showing GIF: {e}")

def stop_gif():
    global gif_running
    gif_running = False

# ---------- START SCREEN ----------

def show_start_screen():
    """Display welcome screen with GIF, name entry, and start button"""
    global child_name, accepting_input
    
    accepting_input = False
    
    # Clear main frame
    for widget in main_frame.winfo_children():
        widget.pack_forget()
    
    # Create start screen container
    start_container = tk.Frame(main_frame, bg="#F4FFDB")
    start_container.pack(expand=True)
    
    # GIF label at top
    gif_label = tk.Label(start_container, bg="#F4FFDB")
    gif_label.pack(pady=20)
    
    # Load and play GIF once
    try:
        gif_path = os.path.join(GIF_PATH, "default_level2.gif")
        gif_image = Image.open(gif_path)
        gif_frames = [
            ImageTk.PhotoImage(frame.resize((600, 400)))
            for frame in ImageSequence.Iterator(gif_image)
        ]
        
        frame_count = [0]  # Use list to modify in nested function
        
        def animate_start_gif():
            if frame_count[0] < len(gif_frames) and gif_label.winfo_exists():
                gif_label.config(image=gif_frames[frame_count[0]])
                gif_label.image = gif_frames[frame_count[0]]
                frame_count[0] += 1
                root.after(100, animate_start_gif)
            elif gif_label.winfo_exists():
                # Keep last frame displayed
                gif_label.config(image=gif_frames[-1])
                gif_label.image = gif_frames[-1]
        
        animate_start_gif()
    except Exception as e:
        print(f"Error loading start GIF: {e}")
    
    # Name entry section
    name_frame = tk.Frame(start_container, bg="#F4FFDB")
    name_frame.pack(pady=30)
    
    name_label = tk.Label(
        name_frame, 
        text="Enter Your Name:", 
        font=("Arial", 32, "bold"),
        bg="#F4FFDB",
        fg="#1d3557"
    )
    name_label.pack(pady=10)
    
    name_entry = tk.Entry(
        name_frame,
        font=("Arial", 28),
        width=20,
        justify="center",
        bg="white",
        fg="#1d3557",
        relief="solid",
        bd=2
    )
    name_entry.pack(pady=10)
    name_entry.focus()
    
    def start_game_clicked():
        global child_name
        child_name = name_entry.get().strip()
        if not child_name:
            child_name = "Player"
        print(f"Starting game for: {child_name}")
        start_container.destroy()
        start_game()
    
    # Bind Enter key
    name_entry.bind("<Return>", lambda e: start_game_clicked())
    
    # Start button
    start_btn = tk.Button(
        name_frame,
        text="START",
        font=("Arial", 32, "bold"),
        bg="#457b9d",
        fg="white",
        activebackground="#1d3557",
        activeforeground="white",
        relief="raised",
        bd=5,
        padx=40,
        pady=15,
        command=start_game_clicked,
        cursor="hand2"
    )
    start_btn.pack(pady=20)

# ---------- GAME FLOW ----------

def start_game():
    global accepting_input, shuffled_animals
    
    # Create shuffled list of all animals for unique questions
    shuffled_animals = animal_list.copy()
    random.shuffle(shuffled_animals)
    
    # Restore game widgets
    image_label.pack(expand=True)
    status_label.pack()
    
    accepting_input = False
    
    # Start directly with first animal (no default GIF)
    show_new_animal()

def show_new_animal():
    global current_animal, retry_count, question_count, accepting_input

    if question_count >= TOTAL_QUESTIONS:
        show_final_score()
        return

    stop_gif()
    retry_count = 0
    
    # Get next unique animal from shuffled list instead of random choice
    current_animal = shuffled_animals[question_count]
    question_count += 1

    print(f"Question {question_count}: {current_animal}")

    try:
        img = Image.open(os.path.join(IMAGE_PATH, f"{current_animal}_image.jpg"))
        img = img.resize((root.winfo_screenwidth(), root.winfo_screenheight()))
        photo = ImageTk.PhotoImage(img)
        image_label.config(image=photo)
        image_label.image = photo
        status_label.config(text=f"Question {question_count}/{TOTAL_QUESTIONS}", fg="#1d3557")
        
        # Enable input after showing the image
        accepting_input = True
    except Exception as e:
        print(f"Error showing animal: {e}")

def check_rfid():
    global retry_count, score, accepting_input

    # Always check for RFID but only process if accepting input
    if ser.in_waiting:
        uid = ser.readline().decode(errors="ignore").strip().upper()
        
        # Ignore if not accepting input or no valid UID
        if not uid or not accepting_input or current_animal is None:
            root.after(100, check_rfid)
            return
            
        print(f"RFID scanned: {uid}")
        
        # Disable input immediately to prevent multiple scans
        accepting_input = False

        if uid == animal_to_uid[current_animal]:
            score += 1
            # Play animal sound first
            play_audio_blocking(os.path.join(SOUND_PATH, f"{current_animal}_sound.mp3"))
            # Then show GIF with correct sound simultaneously
            show_gif_with_audio(
                "goodJob.gif",
                os.path.join(SOUND_PATH, "correct_sound.mp3"),
                duration=2000
            )
            root.after(2500, show_new_animal)

        else:
            retry_count += 1
            if retry_count < MAX_RETRIES:
                # GIF and sound together
                show_gif_with_audio(
                    "tryAgain.gif",
                    os.path.join(SOUND_PATH, "incorrect_sound.mp3"),
                    duration=1500
                )
                # Re-enable input after feedback and showing image again
                def restore_after_retry():
                    show_current_animal()
                    global accepting_input
                    accepting_input = True
                
                root.after(1600, restore_after_retry)
            else:
                # GIF and sound together
                show_gif_with_audio(
                    "uhOh.gif",
                    os.path.join(SOUND_PATH, "oops_sound.mp3"),
                    duration=2000
                )
                root.after(2500, show_new_animal)
    
    root.after(100, check_rfid)

def show_final_score():
    """Display final score with GIF playing once"""
    global gif_running, current_gif_frames, accepting_input
    
    accepting_input = False
    stop_gif()
    
    # Save results to Excel
    save_results_to_excel()
    
    # Hide status label
    status_label.pack_forget()
    image_label.pack_forget()
    
    # Create final screen container
    final_container = tk.Frame(main_frame, bg="#F4FFDB")
    final_container.pack(expand=True)
    
    # GIF label
    final_gif_label = tk.Label(final_container, bg="#F4FFDB")
    final_gif_label.pack(pady=20)
    
    # Score label
    score_text = f"Great Job, {child_name}!\n\nYou identified {score} out of {TOTAL_QUESTIONS} animals correctly!"
    score_label = tk.Label(
        final_container,
        text=score_text,
        font=("Arial", 36, "bold"),
        bg="#F4FFDB",
        fg="#1d3557",
        justify="center"
    )
    score_label.pack(pady=30)
    
    # Load and play final GIF once
    try:
        gif_path = os.path.join(GIF_PATH, "finalScore.gif")
        gif_image = Image.open(gif_path)
        final_frames = [
            ImageTk.PhotoImage(frame.resize((600, 400)))
            for frame in ImageSequence.Iterator(gif_image)
        ]
        
        frame_count = [0]
        
        def animate_final_gif():
            if frame_count[0] < len(final_frames) and final_gif_label.winfo_exists():
                final_gif_label.config(image=final_frames[frame_count[0]])
                final_gif_label.image = final_frames[frame_count[0]]
                frame_count[0] += 1
                root.after(100, animate_final_gif)
            elif final_gif_label.winfo_exists():
                # Keep last frame displayed
                final_gif_label.config(image=final_frames[-1])
                final_gif_label.image = final_frames[-1]
        
        animate_final_gif()
    except Exception as e:
        print(f"Error loading final GIF: {e}")
    
    # Exit button
    exit_btn = tk.Button(
        final_container,
        text="EXIT",
        font=("Arial", 28, "bold"),
        bg="#e63946",
        fg="white",
        activebackground="#d62828",
        activeforeground="white",
        relief="raised",
        bd=5,
        padx=30,
        pady=10,
        command=exit_app,
        cursor="hand2"
    )
    exit_btn.pack(pady=20)

# ---------- START ----------
show_start_screen()
check_rfid()
root.mainloop()
